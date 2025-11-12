#!/usr/bin/env python3
# single_autocall_local_fixed.py
# See docstring inside for details.

from __future__ import annotations

import argparse
import datetime as dt
import json
import os
import re
from typing import Dict, List, Optional, Tuple

import pandas as pd

def parse_date(ds: str) -> Optional[dt.date]:
    from dateutil import parser as dp
    try:
        return dp.parse(ds, fuzzy=True).date()
    except Exception:
        return None

DATE_REGEX = re.compile(r"""(?ix)
\b(?:
    (?:Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?|
     Jul(?:y)?|Aug(?:ust)?|Sep(?:t(?:ember)?)?|Oct(?:ober)?|Nov(?:ember)?|Dec(?:ember)?)
    \s+\d{1,2},\s+\d{4}
  | \d{4}-\d{2}-\d{2}
  | \d{1,2}/\d{1,2}/\d{2,4}
)\b
""")

MONEY_RE = re.compile(r"\$?\s*([0-9]{1,3}(?:[,][0-9]{3})*(?:\.[0-9]+)?|[0-9]+(?:\.[0-9]+)?)")
PCT_RE = re.compile(r"([0-9]+(?:\.[0-9]+)?)\s*%")

def html_to_text(raw: str) -> str:
    try:
        from bs4 import BeautifulSoup
        soup = BeautifulSoup(raw, "lxml")
        return soup.get_text("\n")
    except Exception:
        return raw

def parse_initial_and_threshold(text: str) -> Tuple[Optional[float], Optional[float], Optional[float]]:
    """
    Return (initial, threshold_dollar, threshold_pct_of_initial).
    Prefers the precise dollar amount immediately AFTER 'Downside threshold level'/'Threshold level'.
    Falls back to broader matches; cross-checks % with initial when possible.
    """
    # Initial Share Price
    initial = None
    m_init = re.search(r"initial\s+share\s+price[^$]*\$\s*([0-9,]+(?:\.[0-9]+)?)", text, flags=re.I)
    if m_init:
        initial = float(m_init.group(1).replace(",", ""))

    threshold_dollar: Optional[float] = None
    threshold_pct: Optional[float] = None

    # 1) Prefer the text right AFTER the heading
    for m in re.finditer(r"(downside\s+threshold\s+level|threshold\s+level)", text, flags=re.I):
        start = m.end()                           # look only AFTER the phrase
        end = min(len(text), m.end() + 250)       # tight window to avoid stray examples
        snippet = text[start:end]

        # Prefer high-precision dollar first (e.g., 178.3795)
        m_d = re.search(r"\$?\s*([0-9]{2,5}\.[0-9]{2,5})", snippet)  # bias toward decimals
        if not m_d:
            # fall back to any money format
            m_d = MONEY_RE.search(snippet)

        # Percentage, ideally with "of the initial share price"
        m_p = re.search(r"([0-9]+(?:\.[0-9]+)?)\s*%\s*(?:of\s+the\s+initial\s+share\s+price)?",
                        snippet, flags=re.I)

        if m_d:
            threshold_dollar = float(m_d.group(1).replace(",", ""))
        if m_p:
            threshold_pct = float(m_p.group(1))

        if (threshold_dollar is not None) or (threshold_pct is not None):
            break

    # 2) Wider fallback if nothing found near the heading
    if threshold_dollar is None:
        m_any_d = re.search(r"threshold\s+level[^$]*\$\s*([0-9,]+(?:\.[0-9]+)?)", text, flags=re.I)
        if m_any_d:
            threshold_dollar = float(m_any_d.group(1).replace(",", ""))

    if threshold_pct is None:
        m_any_p = re.search(r"threshold\s+level[^%]*([0-9]+(?:\.[0-9]+)?)\s*%", text, flags=re.I)
        if m_any_p:
            threshold_pct = float(m_any_p.group(1))

    # 3) If only % found and we know initial, compute $
    if threshold_dollar is None and threshold_pct is not None and initial is not None:
        threshold_dollar = round(initial * (threshold_pct / 100.0), 10)

    # 4) Sanity cross-check when we have both $ and %
    if (threshold_dollar is not None) and (threshold_pct is not None) and (initial is not None):
        implied_pct = (threshold_dollar / initial) * 100.0
        # If they disagree a lot (e.g., stray $20), trust the % and recompute $.
        if abs(implied_pct - threshold_pct) > 2.0:
            threshold_dollar = round(initial * (threshold_pct / 100.0), 10)

    return initial, threshold_dollar, threshold_pct


def parse_coupon_apr_and_frequency(text: str):
    coupon_apr = None
    freq = 4
    m_apr = re.search(r"([0-9]+(?:\.[0-9]+)?)\s*%\s*(?:per\s*annum|p\.a\.)", text, flags=re.I)
    if m_apr:
        coupon_apr = float(m_apr.group(1)) / 100.0
    if re.search(r"\bmonthly\b", text, flags=re.I):
        freq = 12
    elif re.search(r"\bquarterly\b", text, flags=re.I):
        freq = 4
    return coupon_apr, freq

def parse_autocall_level(text: str, initial: Optional[float]) -> Optional[float]:
    candidates = []
    for m in re.finditer(r"(automatic(?:ally)?\s+call(?:ed)?|autocall)", text, flags=re.I):
        start = max(0, m.start() - 250)
        end = min(len(text), m.end() + 250)
        candidates.append(text[start:end])
    for m in re.finditer(r"(call\s+level|redemption\s+trigger)", text, flags=re.I):
        start = max(0, m.start() - 250)
        end = min(len(text), m.end() + 250)
        candidates.append(text[start:end])
    level = None
    for s in candidates:
        m_usd = MONEY_RE.search(s)
        m_pct = PCT_RE.search(s)
        if m_usd:
            level = float(m_usd.group(1).replace(",", ""))
            break
        if m_pct and initial is not None:
            level = initial * (float(m_pct.group(1)) / 100.0)
            break
    if level is None and initial is not None and re.search(r"\b100\s*%\s*(?:of\s+the\s+initial|initial\s*share\s*price)", text, flags=re.I):
        level = float(initial)
    return level

def extract_observation_dates_from_tables(html: str) -> List[dt.date]:
    from bs4 import BeautifulSoup
    soup = BeautifulSoup(html, "lxml")
    dates: List[dt.date] = []
    for tbl in soup.find_all("table"):
        rows = []
        for tr in tbl.find_all("tr"):
            cells = [c.get_text(strip=True) for c in tr.find_all(["td","th"])]
            if cells:
                rows.append(cells)
        if not rows:
            continue
        header = rows[0]
        idx = None
        for j,h in enumerate(header):
            if re.search(r"coupon\s+determination\s+date", h, flags=re.I):
                idx = j
                break
        if idx is None:
            continue
        for r in rows[1:]:
            if idx < len(r):
                d = parse_date(r[idx])
                if d:
                    dates.append(d)
    return sorted(set(dates))

def parse_observation_and_final_dates(raw: str) -> List[dt.date]:
    dates: List[dt.date] = []
    try:
        dates = extract_observation_dates_from_tables(raw)
    except Exception:
        dates = []
    if len(dates) < 6:
        text = html_to_text(raw)
        low = text.lower()
        start = low.find("coupon determination dates")
        if start != -1:
            snippet = text[start:start+5000]
            for ds in DATE_REGEX.findall(snippet):
                d = parse_date(ds)
                if d:
                    dates.append(d)
        dates = sorted(set(dates))
    text_all = html_to_text(raw)
    m_final = re.search(r"determination\s+date[^:\n]*[:\-\s]\s*(.+)", text_all, flags=re.I)
    if m_final:
        d = parse_date(m_final.group(1))
        if d:
            dates = sorted(set(dates + [d]))
    return dates

def bulk_download(ticker: str, start: dt.date, end: dt.date) -> pd.DataFrame:
    import yfinance as yf
    print(f"[info] Downloading {ticker} from {start} to {end} (threads=False)…")
    df = yf.download(tickers=ticker, start=start.isoformat(), end=(end+dt.timedelta(days=1)).isoformat(),
                     progress=False, auto_adjust=False, group_by="column", threads=False)
    return df

def close_on_or_prior(df: pd.DataFrame, when: dt.date) -> Optional[float]:
    ser = df["Close"].dropna()
    ser = ser[ser.index.date <= when]
    return float(ser.iloc[-1]) if not ser.empty else None

def build_price_table(dates: List[dt.date], ticker: str, initial: float, threshold: float, autocall_level: Optional[float]) -> pd.DataFrame:
    lo = min(dates) - dt.timedelta(days=14)
    hi = max(dates) + dt.timedelta(days=2)
    df_all = bulk_download(ticker, lo, hi)
    rows = []
    for d in dates:
        print(f"[info] Resolving close on-or-prior to {d} …")
        px = close_on_or_prior(df_all, d)
        row = {"date": d.isoformat(), "ticker": ticker, "close": px, "initial": initial,
               "threshold": threshold, "autocall_level": autocall_level}
        if px is not None:
            row["coupon_ok"] = px >= threshold
            row["autocall_ok"] = (autocall_level is not None) and (px >= autocall_level)
            row["final_ok"] = px >= threshold
            row["ratio_vs_initial"] = px / initial
        else:
            row["coupon_ok"] = None
            row["autocall_ok"] = None
            row["final_ok"] = None
            row["ratio_vs_initial"] = None
        rows.append(row)
    return pd.DataFrame(rows)

def evaluate_single_autocall(price_table: pd.DataFrame, coupon_apr: float, notional: float, payments_per_year: int, first_call_index: int = 1) -> Dict:
    dates = sorted({pd.to_datetime(d).date() for d in price_table["date"].unique()})
    coupon_per_obs = notional * (coupon_apr / payments_per_year)
    schedule = []
    called = False
    call_date = None
    coupons_paid = 0.0
    for idx, d in enumerate(dates):
        sub = price_table[price_table["date"] == d.isoformat()].iloc[0]
        ok_coupon = bool(sub["coupon_ok"]) if pd.notna(sub["coupon_ok"]) else False
        ok_autocall = bool(sub["autocall_ok"]) if pd.notna(sub["autocall_ok"]) else False
        coupon_amt = coupon_per_obs if (ok_coupon and not called) else 0.0
        if coupon_amt:
            coupons_paid += coupon_amt
        status = "Active"
        trigger = False
        if idx >= first_call_index and ok_autocall and not called:
            called = True
            call_date = d
            status = "AutoCalled"
            trigger = True
        row = {"date": d.isoformat(), "coupon_paid": bool(coupon_amt), "coupon_amount": round(coupon_amt, 2),
               "autocall_trigger": trigger, "status": status,
               "close": float(sub["close"]) if pd.notna(sub["close"]) else None}
        schedule.append(row)
        if called:
            break
    if called:
        principal_payoff = notional
        final_ratio = 1.0
    else:
        final_d = pd.to_datetime(schedule[-1]["date"]).date()
        sub = price_table[price_table["date"] == final_d.isoformat()].iloc[0]
        ratio = float(sub["ratio_vs_initial"]) if pd.notna(sub["ratio_vs_initial"]) else None
        final_ok = bool(sub["final_ok"]) if pd.notna(sub["final_ok"]) else False
        if final_ok:
            principal_payoff = notional
        else:
            principal_payoff = notional * (ratio if ratio is not None else 0.0)
        final_ratio = ratio if ratio is not None else None
    total_redemption = principal_payoff + coupons_paid
    profit_loss_pct = ((total_redemption - notional) / notional) * 100.0
    return {"schedule": schedule,
            "summary": {"called": called, "call_date": call_date.isoformat() if call_date else None,
                        "coupons_paid_total": round(coupons_paid, 2),
                        "principal_payoff": round(principal_payoff, 2),
                        "total_redemption": round(total_redemption, 2),
                        "final_ratio": round(final_ratio, 4) if final_ratio is not None else None,
                        "profit_loss_pct": round(profit_loss_pct, 2)}}

def export_to_excel(df_prices: pd.DataFrame, df_sched: pd.DataFrame, summary: Dict, out_path: str = "analysis_output.xlsx"):
    from openpyxl import Workbook
    from openpyxl.chart import LineChart, Reference
    from openpyxl.utils.dataframe import dataframe_to_rows
    wb = Workbook()
    ws_prices = wb.active
    ws_prices.title = "Prices"
    ws_sched = wb.create_sheet("Schedule")
    ws_sum = wb.create_sheet("Summary")
    for r in dataframe_to_rows(df_prices, index=False, header=True):
        ws_prices.append(r)
    if ws_prices.max_row > 2:
        chart = LineChart()
        chart.title = "Close vs Time"
        chart.y_axis.title = "Close"
        chart.x_axis.title = "Observation Index"
        vals = Reference(ws_prices, min_col=3, min_row=1, max_row=ws_prices.max_row)
        chart.add_data(vals, titles_from_data=True)
        ws_prices.add_chart(chart, "L2")
    for r in dataframe_to_rows(df_sched, index=False, header=True):
        ws_sched.append(r)
    for k, v in summary.items():
        ws_sum.append([k, v])
    wb.save(out_path)
    print(f"[done] Excel saved -> {out_path}")

def interactive_select_file(cwd: str) -> str:
    cand = [f for f in os.listdir(cwd) if f.lower().endswith((".html",".htm",".txt"))]
    if not cand:
        raise SystemExit("No .html/.htm/.txt filings found in this folder.")
    print("\nSelect a filing to parse:")
    for i, name in enumerate(cand, 1):
        print(f"  [{i}] {name}")
    while True:
        pick = input("Enter number: ").strip()
        if pick.isdigit() and 1 <= int(pick) <= len(cand):
            return os.path.join(cwd, cand[int(pick)-1])
        print("Invalid selection. Try again.")

def confirm_or_override(dates, initial, threshold_dollar, threshold_pct, autocall_level, coupon_apr, payments_per_year):
    print("\n--- Parsed from filing (confirm or override) ---")
    print(f"Observation dates: {len(dates)} (first/last: {dates[0] if dates else 'N/A'} … {dates[-1] if dates else 'N/A'})")
    print(f"Initial share price: {initial}")
    print(f"Threshold level ($ / % of initial): {threshold_dollar} / {threshold_pct}%")
    print(f"Autocall level ($ or None): {autocall_level}")
    print(f"Coupon APR (decimal): {coupon_apr}")
    print(f"Payments per year: {payments_per_year}")
    yn = input("Proceed with these? [Y/n]: ").strip().lower()
    if yn == "n":
        ds = input("Enter observation dates (comma-separated, YYYY-MM-DD): ").strip()
        if ds:
            nd = [parse_date(d) for d in ds.split(",")]
            dates = [d for d in nd if d]
            dates.sort()
        def need_float(prompt: str, default: Optional[float]) -> float:
            while True:
                val = input(f"{prompt} [{default if default is not None else ''}]: ").strip()
                if not val and default is not None:
                    return float(default)
                try:
                    return float(val)
                except Exception:
                    print("Invalid number, try again.")
        initial = need_float("Initial share price", initial)
        threshold_dollar = need_float("Threshold level ($)", threshold_dollar)
        ac = input(f"Autocall level ($) or blank for None [{autocall_level if autocall_level is not None else ''}]: ").strip()
        autocall_level = float(ac) if ac else None
        coupon_apr = need_float("Coupon APR (decimal, e.g., 0.1065 for 10.65%)", coupon_apr)
        freq_in = input(f"Payments per year [monthly=12, quarterly=4] [{payments_per_year}]: ").strip()
        if freq_in:
            try:
                payments_per_year = int(freq_in)
                if payments_per_year not in (4,12):
                    print("Invalid frequency; defaulting to 4 (quarterly).")
                    payments_per_year = 4
            except Exception:
                print("Invalid input; keeping previous value.")
    if not dates:
        raise SystemExit("No observation dates provided.")
    if initial is None or threshold_dollar is None or coupon_apr is None:
        raise SystemExit("Initial, Threshold ($), and Coupon APR are required.")
    return dates, float(initial), float(threshold_dollar), autocall_level, float(coupon_apr), int(payments_per_year)

def main():
    ap = argparse.ArgumentParser(description="Single-underlying autocallable note analyzer (LOCAL files only)")
    ap.add_argument("--excel", action="store_true", help="Also write a single Excel workbook.")
    ap.add_argument("--out-prefix", help="Override output file prefix.")
    ap.add_argument("--notional", type=float, default=1000.0)
    ap.add_argument("--first-call-index", type=int, default=1, help="Index (0-based) of first date eligible for autocall; 1 means 2nd obs date.")
    args = ap.parse_args()
    source = interactive_select_file(os.getcwd())
    print(f"\n[info] Using: {source}")
    raw = open(source, "r", encoding="utf-8", errors="ignore").read()
    dates = parse_observation_and_final_dates(raw)
    text_all = html_to_text(raw)
    initial, threshold_dollar, threshold_pct = parse_initial_and_threshold(text_all)
    autocall_level = parse_autocall_level(text_all, initial)
    coupon_apr_parsed, freq_parsed = parse_coupon_apr_and_frequency(text_all)
    payments_per_year = freq_parsed or 4
    coupon_apr = coupon_apr_parsed
    dates, initial, threshold_dollar, autocall_level, coupon_apr, payments_per_year = confirm_or_override(
        dates, initial, threshold_dollar, threshold_pct, autocall_level, coupon_apr, payments_per_year)
    ticker = (input("Enter underlying ticker (e.g., DOCU): ").strip() or "TICKER").upper()
    def nice_prefix(ticker: str, dates: List[dt.date]) -> str:
        if dates:
            return f"{ticker}_{dates[0].isoformat()}_to_{dates[-1].isoformat()}"
        return ticker
    prefix = args.out_prefix or nice_prefix(ticker, dates)
    df_prices = build_price_table(dates, ticker, initial, threshold_dollar, autocall_level)
    result = evaluate_single_autocall(df_prices, coupon_apr, args.notional, payments_per_year, args.first_call_index)
    df_sched = pd.DataFrame(result["schedule"])
    summary = result["summary"]
    prices_csv = f"{prefix}_prices.csv"
    sched_csv = f"{prefix}_schedule.csv"
    summary_json = f"{prefix}_summary.json"
    df_prices.to_csv(prices_csv, index=False)
    df_sched.to_csv(sched_csv, index=False)
    with open(summary_json, "w", encoding="utf-8") as f:
        json.dump(summary, f, indent=2)
    print(f"\n[done] Wrote:\n  {prices_csv}\n  {sched_csv}\n  {summary_json}")
    if args.excel:
        xlsx = f"{prefix}.xlsx"
        export_to_excel(df_prices, df_sched, summary, xlsx)
    print("\n=== PRICE TABLE (first 12 rows) ===")
    print(df_prices.head(12).to_string(index=False))
    print("\n=== SUMMARY ===")
    print(pd.Series(summary).to_string())

if __name__ == "__main__":
    main()
